# -*- coding: utf-8 -*-
"""
Заполнение листа «Опросник ДКБ» — ответы из фактов проекта WorldCashFit.
Колонка E: что реализовано + что заказчик приложит документами.
"""
from __future__ import annotations

import shutil
from pathlib import Path

import openpyxl

SRC = Path(r"c:\Users\anton\Downloads\Карточка партнера.xlsx")
OUT = Path(r"c:\Users\anton\Downloads\Карточка партнера_заполнено.xlsx")

COL_SCORE = 4
COL_EXPLANATION = 5

PRIVACY = "https://worldcashfit.ru/privacy"
LEGAL = "https://worldcashfit.ru/legal"
CONSENT = "https://worldcashfit.ru/personal-data-consent"
OFFER = "https://worldcashfit.ru/license_agreement/"

ARCH = (
    "Схема: мобильные приложения WorldCashFit (iOS App Store, Android Google Play ru.worldcashfit.app) "
    "и браузер CRM → HTTPS (TLS 1.2+, Let's Encrypt, nginx на worldcashfit.ru:443) → "
    "backend Symfony 7 / PHP 8.4 (Docker-контейнер app:8000) → MySQL 8.0 (Docker-контейнер database, "
    "сеть crm_backend_symfony_net; порт 3306 НЕ опубликован в интернет). "
    "Обмен с ПАО Сбербанк: oauth.sber.ru / id.sber.ru по mTLS (PKCS#12), OAuth 2.0 + PKCE, "
    "callback https://worldcashfit.ru/api/v1/auth/sber/callback."
)

PD_SBER = (
    "ПДн от Банка (пакет Professional, scope maindoc): ФИО, дата рождения, пол, телефон, e-mail, "
    "паспорт РФ (серия, номер, кем выдан, дата выдачи, код подразделения) — "
    "получаются только через GET userinfo после явного согласия пользователя в Сбер ID."
)

TECH_ACCESS = (
    "Доступ к API клиентов — только Bearer JWT (access-token, TTL 3600 с, MOBILE_API_ACCESS_TTL). "
    "Заголовок X-User-Id отключён (исключена подмена пользователя). "
    "CRM /admin — form_login + CSRF, роли ROLE_STAFF/ROLE_ADMIN/ROLE_SUPER_ADMIN. "
    "Паспорт в CRM видят только ROLE_ADMIN и ROLE_SUPER_ADMIN (PassportAccessPolicy). "
    "После верификации через Сбер ID клиент не может изменить паспорт через API (isPassportLockedFromClientEdit). "
    "Блокировка клиента: isBlocked → 401 в приложении. "
    "Удаление аккаунта: DELETE /api/v1/user/account — обезличивание e-mail, ФИО, телефона, паспорта, sber_id, токенов."
)

CLIENT_DOCS = "Заказчику приложить к анкете:"

ANSWERS: dict[int, tuple[str, bool]] = {
    1: (
        f"Оператор ПДн: ООО «Ворлдкэшбокс», ИНН 2543082240, ОГРН 1152543020629, "
        f"690014, Приморский край, г. Владивосток, ул. Толстого, 32А, оф. 308. "
        f"Реестр операторов ПДн: https://pd.rkn.gov.ru/operators-registry/operators-list/ "
        f"(поиск по ИНН 2543082240).\n"
        f"{CLIENT_DOCS} регистрационный номер в реестре РКН и прямую URL-ссылку на карточку оператора "
        f"(из личного кабинета / результата поиска на pd.rkn.gov.ru).",
        True,
    ),
    2: (
        f"Политика обработки ПДн размещена публично: {PRIVACY} "
        f"(также раздел {LEGAL}). "
        f"Согласие на обработку ПДn: {CONSENT}. "
        f"Оферта: {OFFER}.\n"
        f"{CLIENT_DOCS} скан-копию утверждённой Политики с подписью руководителя (если Сбер требует бумажный экземпляр помимо URL).",
        True,
    ),
    3: (
        f"{PD_SBER} ИС: CRM/API WorldCashFit, сайт worldcashfit.ru, приложения iOS/Android.\n"
        f"{CLIENT_DOCS} акт / приказ «Оценка вреда, который может быть причинён субъектам ПДн» "
        f"(скан с реквизитами, датой утверждения).",
        True,
    ),
    4: (
        f"ИСПДн, в которых обрабатываются ПДн от Банка: (1) backend https://worldcashfit.ru/api/v1/ "
        f"(Symfony + MySQL); (2) админ-панель https://worldcashfit.ru/admin; (3) мобильные клиенты iOS/Android. "
        f"Технические меры: TLS, RBAC, JWT, mTLS к Сберу, изоляция БД в Docker-сети.\n"
        f"{CLIENT_DOCS} приказ / акт об определении уровня защищённости ИСПДn (рекомендуемый уровень для данной ИС — УЗ-2).",
        True,
    ),
    5: (
        f"БД персональных данных: MySQL 8.0, Docker volume database_data на VPS в РФ. "
        f"Порт 3306 не публикуется на 0.0.0.0 (compose.yaml — только внутренняя сеть crm_backend_symfony_net). "
        f"Запись, хранение, уточнение, извлечение, удаление — через backend на территории РФ. "
        f"{ARCH}\n"
        f"{CLIENT_DOCS} договор аренды VPS/хостинга с указанием локализации серверов в РФ; "
        f"при необходимости — описание процесса обработки (можно приложить распечатку схемы из п. 11/26 настоящей анкеты).",
        True,
    ),
    6: (
        f"В системе реализовано уничтожение/обезличивание по запросу пользователя: "
        f"DELETE /api/v1/user/account (UserAccountDeletionService) — обнуляются/анonymizируются "
        f"e-mail, ФИО, телефон, дата/место рождения, все поля паспорта, адрес регистрации, sber_id, токены; "
        f"аккаунт блокируется. Паспорт после Сбер ID не редактируется клиентом через API.\n"
        f"{CLIENT_DOCS} локальный акт «Порядок уточнения, блокирования, прекращения обработки и уничтожения ПДн» "
        f"(приказ или положение с реквизитами).",
        True,
    ),
    7: (
        f"Backend размещён на Linux VPS; антивирус на уровне сервера — зона ответственности эксплуатации.\n"
        f"{CLIENT_DOCS} лицензии/скриншоты консоли антивируса на VPS и рабочих станциях администраторов "
        f"(Kaspersky / Dr.Web / иное — что фактически используется).",
        True,
    ),
    8: (
        f"Обновления безопасности ОС и Docker-образов применяются администратором сервера при эксплуатации VPS.\n"
        f"{CLIENT_DOCS} скриншоты актуальных обновлений на VPS на текущую дату "
        f"(uname -a, apt list --upgradable или аналог, версия Docker).",
        True,
    ),
    9: (
        f"Технически: логирование auth-ошибок (401), логи nginx (access/error), логи Symfony (var/log); "
        f"лог userinfo Сбер ID по умолчанию выключен (SBER_USERINFO_LOG=0), блок identification в лог не пишется.\n"
        f"{CLIENT_DOCS} регламент реагирования на инциденты КБ и журнал учёта инцидентов (реквизиты приказа + выдержка).",
        True,
    ),
    10: (
        f"На VPS: firewall — открыты только 22 (SSH), 80→301 HTTPS, 443. "
        f"MySQL недоступен из интернета. Единая точка входа HTTP(S) — nginx (docker/nginx/default.conf). "
        f"Межконтейнерный трафик — изолированная Docker-сеть.\n"
        f"{CLIENT_DOCS} положение / регламент межсетевого экранирования (реквизиты утверждения).",
        True,
    ),
    11: (
        f"{ARCH} "
        f"Дополнительно: шлюз турникета в клубе обращается к CRM по Authorization: Bearer <gateway_token> "
        f"(отдельный токен клуба, не публикуется в приложении клиента).\n"
        f"{CLIENT_DOCS} сетевая схема в PDF/Visio (можно оформить на основе текста выше).",
        True,
    ),
    12: (
        f"Удалённое администрирование production-сервера — SSH по ключу (не по паролю). "
        f"Доступ к CRM — HTTPS через браузер, учётные записи StaffUser с паролем (bcrypt) и CSRF.\n"
        f"{CLIENT_DOCS} регламент удалённого доступа (SSH/VPN, кто имеет доступ, как выдаётся/отзывается).",
        True,
    ),
    13: (
        f"Контакт по вопросам ПДn в продукте: hello@worldcashfit.ru, vl@worldcashbox.ru, тел. 8 994 010 72 72.\n"
        f"{CLIENT_DOCS} приказ о назначении лица, ответственного за организацию обработки ПДn "
        f"(ФИО, должность, дата, номер приказа).",
        True,
    ),
    14: (
        f"Публичные документы на сайте: политика {PRIVACY}, согласие https://worldcashfit.ru/personal-data-consent, "
        f"оферта https://worldcashfit.ru/license_agreement/, договоры с клиентом/тренером — в разделе legal.\n"
        f"{TECH_ACCESS}\n"
        f"{CLIENT_DOCS} приказы об утверждении Политики ПДn, Положение об обработке ПДn, "
        f"локальные акты по предотвращению нарушений 152-ФЗ (скан-копии с реквизитами).",
        True,
    ),
    15: (
        f"{TECH_ACCESS} "
        f"Меню CRM разграничено по ролям (AdminMenuBuilder): crm_staff и franchise — только SUPER_ADMIN/ADMIN; "
        f"остальные разделы — по ROLE_STAFF и специализациям (manager, finance и т.д.).\n"
        f"{CLIENT_DOCS} «Положение о разграничении доступа к ИСПДn» / «Правила доступа к CRM WorldCashFit».",
        True,
    ),
    16: (
        f"Разработчики и администраторы CRM работают с 152-ФЗ и политикой на сайте; "
        f"техническая документация по интеграции Сбер ID — etc/sber/README.md (внутри проекта).\n"
        f"{CLIENT_DOCS} листы ознакомления сотрудников с Политикой ПДn и 152-ФЗ (журнал / подписи).",
        True,
    ),
    17: (
        f"{CLIENT_DOCS} подписанные обязательства о неразглашении КИ и ПДn с каждым сотрудником, "
        f"имеющим доступ к CRM (скан, ПДn работника затушевать).",
        True,
    ),
    18: (
        f"{CLIENT_DOCS} должностные инструкции администратора CRM / системного администратора "
        f"с разделом об обработке и защите ПДn (скан, ПДn затушевать).",
        True,
    ),
    19: (
        f"Технические меры против угроз: HTTPS, mTLS к Сберу, PKCE+state+nonce, JWT с истечением, "
        f"запрет подмены user id, блокировка аккаунта, маскирование паспорта в логах, "
        f"изоляция БД, CSRF в admin, хеширование паролей staff (Symfony password_hasher auto).\n"
        f"{CLIENT_DOCS} «Модель угроз» / акт определения актуальных угроз для ИСПДn WorldCashFit.",
        True,
    ),
    20: (
        f"Контроль на уровне эксплуатации: code review, git, деплой через docker compose build; "
        f"миграции БД doctrine:migrations; мониторинг доступности сервиса.\n"
        f"{CLIENT_DOCS} план внутреннего контроля / регламент периодической проверки мер защиты ИСПДn.",
        True,
    ),
    21: (
        f"Production-ИС размещена на VPS/DC в РФ (физическая охрана — у хостинг-провайдера). "
        f"Локальные серверные помещения заказчика для этой ИС не используются.\n"
        f"{CLIENT_DOCS} при необходимости — выписка из договора с DC о физической защите / SLA хостинга.",
        True,
    ),
    22: (
        f"Организационные и технические меры (реализовано в продукте WorldCashFit):\n"
        f"• {ARCH}\n"
        f"• {PD_SBER}\n"
        f"• {TECH_ACCESS}\n"
        f"• OAuth 2.0 Authorization Code + PKCE (SberOAuthPkceStateService, SberIdTokenValidator)\n"
        f"• mTLS PKCS#12 к API Сбера (SberIdOAuthService, сертификат в etc/sber/)\n"
        f"• Scope OIDC: openid profile email mobile name birthdate gender maindoc priority_doc\n"
        f"• Лог userinfo: выкл. по умолчанию; identification → [REDACTED] (SberIdUserinfoJsonLogger)\n"
        f"• Android: HTTP body в лог только DEBUG-сборка (BuildConfig.DEBUG); prod — без логирования тел запросов\n"
        f"• network_security_config.xml — только HTTPS к worldcashfit.ru\n"
        f"{CLIENT_DOCS} приказ «О мерах по обеспечению безопасности ПДn в ИС WorldCashFit» (можно приложить распечатку текста этого пункта + подпись руководителя).",
        True,
    ),
    23: (
        f"Техническое сопровождение и администрирование сервера — команда разработки/эксплуатации WorldCashFit; "
        f"контакт hello@worldcashfit.ru.\n"
        f"{CLIENT_DOCS} приказ о назначении лица, ответственного за информационную безопасность "
        f"(ФИО, должность, функции, права).",
        True,
    ),
    24: (
        f"В системе: пароли StaffUser — Symfony password_hasher (bcrypt/argon auto); "
        f"CSRF на форме входа /admin; JWT access 3600 с; refresh-token в БД; "
        f"gateway_token и ACCESS_GATE_TOKEN — отдельные секреты для шлюза турникета.\n"
        f"{CLIENT_DOCS} документ «Парольная политика» (мин. длина, срок смены, запрет передачи).",
        True,
    ),
    25: (
        f"Перечень ВНД, которые заказчик утверждает для организации (шаблон):\n"
        f"1. Политика обработки ПДn (URL: {PRIVACY})\n"
        f"2. Положение об информационной безопасности\n"
        f"3. Регламент резервного копирования\n"
        f"4. Регламент управления доступом к CRM\n"
        f"5. Регламент реагирования на инциденты ИБ\n"
        f"6. Парольная политика\n"
        f"{CLIENT_DOCS} реестр утверждённых ВНД с датами и номерами приказов.",
        True,
    ),
    26: (
        f"Архитектура ИТ-инфраструктуры WorldCashFit (production):\n"
        f"1. Клиентский канал: iOS-приложение (App Store) и Android-приложение (Google Play, package ru.worldcashfit.app) "
        f"→ HTTPS → worldcashfit.ru:443 (nginx, TLS Let's Encrypt).\n"
        f"2. nginx (контейнер crm_backend_nginx, compose.https.yaml) проксирует на Symfony app:8000.\n"
        f"3. Backend: Symfony 7, PHP 8.4, Docker-образ из Dockerfile; маршруты /api/v1/*, /admin.\n"
        f"4. БД: MySQL 8.0, контейнер crm_backend_db, volume database_data, только docker-сеть.\n"
        f"5. Интеграция Сбер ID: app → mTLS → oauth.sber.ru / id.sber.ru; callback GET/POST /api/v1/auth/sber/callback.\n"
        f"6. CRM staff: браузер → HTTPS → /admin (ROLE_STAFF+, CSRF).\n"
        f"7. Турникет (опционально): LAN-шлюз → Bearer gateway_token → /api/v1/gateway/access/entry.\n"
        f"{CLIENT_DOCS} PDF/Visio-схему на основе описания выше (оформление — у заказчика/ИБ).",
        True,
    ),
    27: (
        f"Централизованная MDM/BYOD-система для личных устройств сотрудников в коде продукта не предусмотрена. "
        f"Доступ к CRM — с авторизованных рабочих мест через HTTPS; мобильные клиенты — только официальные приложения из Store.\n"
        f"{CLIENT_DOCS} положение об использовании личных устройств (или запрет BYOD для доступа к CRM).",
        False,
    ),
    28: (
        f"DLP-система не развёрнута. Компенсирующие технические меры в продукте: "
        f"экспорт/просмотр паспорта только ROLE_ADMIN; паспорт не логируется (SBER_USERINFO_LOG=0, redact identification); "
        f"передача только по TLS; клиент не может выгрузить чужие данные (JWT per user).\n"
        f"{CLIENT_DOCS} скриншоты настроек nginx/TLS и CRM (роли без доступа к паспорту для manager) — "
        f"либо обоснование отказа от DLP для облачной ИС малого масштаба.",
        False,
    ),
    29: (
        f"На сервере (VPS): Linux, Docker, nginx, PHP 8.4, MySQL 8, git. "
        f"На АРМ администраторов: браузер, SSH-клиент. Запрещено: торрент-клиенты, нелицензионное ПО.\n"
        f"{CLIENT_DOCS} утверждённый «Перечень разрешённого и запрещённого ПО».",
        True,
    ),
    30: (
        f"SSH-доступ к VPS — по ключу; CRM — HTTPS + индивидуальные учётки StaffUser; "
        f"API — только с валидным Bearer-токеном.\n"
        f"{CLIENT_DOCS} выдержка из регламента: удалённый доступ только с устройств "
        f"с актуальной ОС, антивирусом и блокировкой экрана.",
        True,
    ),
    31: (
        f"Обработка ПДn на production-серверах WorldCashFit выполняется по проводной инфраструктуре VPS/DC, не по Wi-Fi. "
        f"Wi-Fi офиса заказчика к серверу обработки ПДn не подключён.\n"
        f"{CLIENT_DOCS} при наличии офисной Wi-Fi — политика WPA2/WPA3 и скрин настроек точки доступа.",
        True,
    ),
}


def fill_dkb_sheet(wb: openpyxl.Workbook, orig_wb: openpyxl.Workbook) -> None:
    ws = wb["Опросник ДКБ"]
    orig = orig_wb["Опросник ДКБ"]

    for row in range(7, 39):
        num = ws.cell(row=row, column=2).value
        if num is None or not isinstance(num, (int, float)):
            continue
        n = int(num)
        if n not in ANSWERS:
            continue
        text, applicable = ANSWERS[n]
        ws.cell(row=row, column=COL_SCORE, value="Да" if applicable else "НП")
        ws.cell(row=row, column=COL_EXPLANATION, value=text)
        for col in (6, 7):
            ws.cell(row=row, column=col, value=orig.cell(row=row, column=col).value)

        # перенос текста в ячейке
        ws.cell(row=row, column=COL_EXPLANATION).alignment = openpyxl.styles.Alignment(
            wrap_text=True, vertical="top"
        )


def main() -> None:
    orig_wb = openpyxl.load_workbook(SRC)
    shutil.copy2(SRC, OUT)
    wb = openpyxl.load_workbook(OUT)
    fill_dkb_sheet(wb, orig_wb)
    wb.save(OUT)
    print(f"Saved: {OUT}")


if __name__ == "__main__":
    main()
